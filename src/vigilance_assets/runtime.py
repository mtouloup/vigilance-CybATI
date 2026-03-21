from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook

from .config import (
    FILE_BACKEND,
    GOOGLE_SHEETS_BACKEND,
    MEMORY_BACKEND,
    AppRuntimeSettings,
    ConfigurationError,
    load_runtime_settings,
)
from .repository import SpreadsheetAssetRepository
from .spreadsheet import SheetRecord, SpreadsheetTableGateway


@dataclass(slots=True)
class InMemorySpreadsheetGateway(SpreadsheetTableGateway):
    """Minimal header-keyed gateway used for tests and default local startup."""

    _rows_by_sheet: dict[str, list[SheetRecord]] = field(default_factory=dict)

    def list_rows(self, sheet_name: str) -> list[SheetRecord]:
        return [
            SheetRecord(row_number=row.row_number, values=dict(row.values))
            for row in self._rows_by_sheet.get(sheet_name, [])
        ]

    def append_row(self, sheet_name: str, values: dict[str, object]) -> SheetRecord:
        rows = self._rows_by_sheet.setdefault(sheet_name, [])
        record = SheetRecord(row_number=len(rows) + 2, values=dict(values))
        rows.append(record)
        return SheetRecord(row_number=record.row_number, values=dict(record.values))

    def update_row(self, sheet_name: str, row_number: int, values: dict[str, object]) -> SheetRecord:
        rows = self._rows_by_sheet.setdefault(sheet_name, [])
        for index, record in enumerate(rows):
            if record.row_number == row_number:
                updated = SheetRecord(row_number=row_number, values=dict(values))
                rows[index] = updated
                return SheetRecord(row_number=updated.row_number, values=dict(updated.values))
        raise KeyError(f"Row {row_number} not found in sheet {sheet_name}.")

    def delete_row(self, sheet_name: str, row_number: int) -> None:
        rows = self._rows_by_sheet.setdefault(sheet_name, [])
        for index, record in enumerate(rows):
            if record.row_number == row_number:
                del rows[index]
                return
        raise KeyError(f"Row {row_number} not found in sheet {sheet_name}.")


@dataclass(slots=True)
class WorkbookFileGateway(SpreadsheetTableGateway):
    """openpyxl-backed gateway for local file runtime usage."""

    path: Path
    assets_sheet_name: str = "ASSETS"
    vocabularies_sheet_name: str = "VOCABULARIES"
    expected_headers: Sequence[str] = ()
    read_only: bool = False

    def list_rows(self, sheet_name: str) -> list[SheetRecord]:
        workbook = self._open_workbook()
        sheet = self._get_sheet(workbook, self._resolve_sheet_name(sheet_name))
        header_row = self._header_row(sheet)
        headers = [self._normalize_cell(cell.value) for cell in sheet[header_row]]
        records: list[SheetRecord] = []
        for row_index in range(header_row + 1, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_index, column=column).value for column in range(1, len(headers) + 1)]
            normalized_values = [self._normalize_cell(value) for value in row_values]
            if not any(value not in (None, "") for value in normalized_values):
                continue
            records.append(
                SheetRecord(
                    row_number=row_index,
                    values={header: normalized_values[index] for index, header in enumerate(headers) if header},
                )
            )
        workbook.close()
        return records

    def append_row(self, sheet_name: str, values: dict[str, object]) -> SheetRecord:
        self._ensure_writable()
        workbook = self._open_workbook()
        sheet = self._ensure_sheet(workbook, self._resolve_sheet_name(sheet_name))
        headers = self._ensure_headers(sheet)
        row_number = sheet.max_row + 1
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_number, column=column_index, value=values.get(header))
        self._save_workbook(workbook)
        workbook.close()
        return SheetRecord(row_number=row_number, values={header: values.get(header) for header in headers})

    def update_row(self, sheet_name: str, row_number: int, values: dict[str, object]) -> SheetRecord:
        self._ensure_writable()
        workbook = self._open_workbook()
        sheet = self._get_sheet(workbook, self._resolve_sheet_name(sheet_name))
        headers = self._ensure_headers(sheet)
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_number, column=column_index, value=values.get(header))
        self._save_workbook(workbook)
        workbook.close()
        return SheetRecord(row_number=row_number, values={header: values.get(header) for header in headers})

    def delete_row(self, sheet_name: str, row_number: int) -> None:
        self._ensure_writable()
        workbook = self._open_workbook()
        sheet = self._get_sheet(workbook, self._resolve_sheet_name(sheet_name))
        sheet.delete_rows(row_number, 1)
        self._save_workbook(workbook)
        workbook.close()

    def _resolve_sheet_name(self, sheet_name: str) -> str:
        if sheet_name == "ASSETS":
            return self.assets_sheet_name
        if sheet_name == "VOCABULARIES":
            return self.vocabularies_sheet_name
        return sheet_name

    def _open_workbook(self):
        if self.path.exists():
            return load_workbook(self.path)
        workbook = Workbook()
        default_sheet = workbook.active
        default_sheet.title = self.assets_sheet_name
        self._ensure_headers(default_sheet)
        workbook.create_sheet(self.vocabularies_sheet_name)
        self._save_workbook(workbook)
        return workbook

    def _ensure_sheet(self, workbook, sheet_name: str):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        return workbook.create_sheet(sheet_name)

    def _get_sheet(self, workbook, sheet_name: str):
        if sheet_name not in workbook.sheetnames:
            raise ConfigurationError(f"Workbook {self.path} does not contain worksheet {sheet_name!r}.")
        return workbook[sheet_name]

    def _header_row(self, sheet) -> int:
        return 1

    def _ensure_headers(self, sheet) -> tuple[str, ...]:
        existing = (
            [self._normalize_cell(cell.value) for cell in sheet[1] if self._normalize_cell(cell.value)]
            if sheet.max_row >= 1
            else []
        )
        headers = tuple(existing or self.expected_headers)
        if not headers:
            raise ConfigurationError("Workbook file backend requires expected headers to initialize the ASSETS sheet.")
        if not existing:
            for column_index, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=column_index, value=header)
        return headers

    def _save_workbook(self, workbook) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(self.path)

    def _ensure_writable(self) -> None:
        if self.read_only:
            raise ConfigurationError(f"Workbook file backend at {self.path} is read-only.")

    @staticmethod
    def _normalize_cell(value: object) -> object:
        if value is None:
            return None
        if isinstance(value, str):
            stripped = value.strip()
            return stripped or None
        return value


def create_repository_from_settings(settings: AppRuntimeSettings) -> SpreadsheetAssetRepository:
    from .google_sheets import build_google_sheets_gateway
    from .spreadsheet import AssetSpreadsheetMapper

    mapper = AssetSpreadsheetMapper()
    backend = settings.spreadsheet.backend
    if backend == MEMORY_BACKEND:
        gateway = InMemorySpreadsheetGateway()
    elif backend == FILE_BACKEND:
        if not settings.spreadsheet.file.path:
            raise ConfigurationError("File backend selected without VIGILANCE_SPREADSHEET_FILE_PATH.")
        gateway = WorkbookFileGateway(
            path=Path(settings.spreadsheet.file.path),
            assets_sheet_name=settings.spreadsheet.sheets.assets,
            vocabularies_sheet_name=settings.spreadsheet.sheets.vocabularies,
            expected_headers=mapper.ordered_headers,
            read_only=settings.spreadsheet.file.read_only,
        )
        gateway.list_rows("ASSETS")
    elif backend == GOOGLE_SHEETS_BACKEND:
        gateway = build_google_sheets_gateway(settings.spreadsheet)
    else:
        raise ConfigurationError(f"Unsupported spreadsheet backend: {backend}")

    return SpreadsheetAssetRepository(
        workbook_reference=settings.spreadsheet.workbook_reference,
        gateway=gateway,
    )


def create_runtime_app():
    from .api import create_app

    settings = load_runtime_settings()
    repository = create_repository_from_settings(settings)
    app = create_app(repository=repository)
    app.config["RUNTIME_SETTINGS"] = settings
    return app
